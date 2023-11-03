from openpyxl import load_workbook

#la position du premier nom de joueur dans le tableau de disponibilités
disposXshift =4
disposYshift = 5

#la position du premier nom de joueur dans le tableau de progression
progressXshift = 25
progressYshift = 5



def getData():
    data = load_workbook("planning.xlsx")
    sheet = data["dispo"]

    nbPlayers = int(sheet.cell(2,1).value)
    nbSlots = int(sheet.cell(2,2).value)
    nbMissions = int(sheet.cell(2,3).value)

    x = disposXshift
    y = disposYshift
    players = []
    dispos = {}
    while len(players) < nbPlayers:
        name = sheet.cell(y,x).value
        if name != None:
            players += [name]
            dispos[name] = []
            for slotIndex in range(nbSlots):
                value = sheet.cell(y,x+1+slotIndex).value
                dispos[name] += [1 if value != None else 0]
        y += 1

    x = progressXshift
    y = progressYshift
    skipper = {}
    progress = {}
    finished = False
    while not finished:
        name = sheet.cell(y,x).value
        if name != None:
            if name in players:
                skipperValue = sheet.cell(y,x+1).value
                skipper[name] = int(skipperValue) if skipperValue != None else 0

                progress[name] = []
                for missionIndex in range(nbMissions):
                    value = sheet.cell(y,x+2+missionIndex).value
                    progress[name] += [1 if value != None else 0]
        else:
            finished = True
        y += 1

    slotsDates = []
    x = disposXshift+1
    y = disposYshift-1
    for slotIndex in range(nbSlots):
        slotDate = sheet.cell(y,x+slotIndex).value + " "
        slotDate += sheet.cell(y-1,x+slotIndex).value
        slotsDates += [slotDate]

    instance = [nbPlayers,nbSlots,nbMissions,players,slotsDates,dispos,skipper,progress]
    return instance


def allowSkipping(instance):
    nbPlayers,nbSlots,nbMissions,players,slotsDates,dispos,skipper,progress = instance

    newProgress = {}
    for player in players:
        newProgress[player] = [0]*nbMissions
        skipping = False
        for missionIndex in range(nbMissions):
            newProgress[player][missionIndex] = progress[player][missionIndex]

            if skipping:
                newProgress[player][missionIndex] = 1

            if skipping == False and skipper[player] == 1 and progress[player][missionIndex] == 1:
                skipping = True

    return newProgress


def applySkipping(instance):
    newProgress = allowSkipping(instance)
    nbPlayers,nbSlots,nbMissions,players,slotsDates,dispos,skipper,progress = instance
    instance = [nbPlayers,nbSlots,nbMissions,players,slotsDates,dispos,skipper,newProgress]
    return instance
